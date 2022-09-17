from dataclasses import dataclass
import openpyxl


@dataclass(frozen=True, eq=True)
class DayShift:
    day: str
    shift: str


@dataclass(frozen=True, eq=True)
class PostingDemand:
    echo: int
    irm: int
    tdm: int


EXCEL_DAYSHIFT_OFFSET = 5
ALL_DAYSHIFTS = [
    DayShift("LUNDI", "MATIN"),
    DayShift("LUNDI", "APRES-MIDI"),
    DayShift("MARDI", "MATIN"),
    DayShift("MARDI", "APRES-MIDI"),
    DayShift("MERCREDI", "MATIN"),
    DayShift("MERCREDI", "APRES-MIDI"),
    DayShift("JEUDI", "MATIN"),
    DayShift("JEUDI", "APRES-MIDI"),
    DayShift("VENDREDI", "MATIN"),
    DayShift("VENDREDI", "APRES-MIDI"),
]


@dataclass(frozen=True, eq=True)
class Agent:
    name: str
    echo_capacity: int
    irm_capacity: int
    tdm_capacity: int
    off: list[DayShift]


@dataclass(frozen=True, eq=True)
class Configuration:
    staffing_minimum: bool
    vacation_maximum: bool
    agents: list[Agent]
    demands: dict[DayShift, PostingDemand]


def str_to_bool(v: str, ctx_message: str = None) -> bool:
    if v == "VRAI":
        return True
    elif v == "FAUX":
        return False
    raise ValueError(
        f'Error while reading {ctx_message}- Expected "VRAI" or "FAUX", but got {v}'
    )


def read_configuration() -> Configuration:
    workbook = openpyxl.load_workbook("configuration.xlsx")
    worksheet = workbook.active

    # Read initial configuration
    staffing_minimum = str_to_bool(worksheet.cell(3, 5).value, "STAFFING MINIMUM (E3)")
    vacation_maximum = str_to_bool(worksheet.cell(3, 6).value, "VACATION_MAXIMUM (F3)")

    # Read the demands for each posting
    ECHO_ROW = 6
    IRM_ROW = 7
    TDM_ROW = 8
    demands = {}
    for i in range(0, len(ALL_DAYSHIFTS)):
        echo = worksheet.cell(ECHO_ROW, i + EXCEL_DAYSHIFT_OFFSET).value
        irm = worksheet.cell(IRM_ROW, i + EXCEL_DAYSHIFT_OFFSET).value
        tdm = worksheet.cell(TDM_ROW, i + EXCEL_DAYSHIFT_OFFSET).value
        demands[ALL_DAYSHIFTS[i]] = PostingDemand(echo=echo, irm=irm, tdm=tdm)

    # Read the agents
    AGENT_ROW_START = 11
    current_row = AGENT_ROW_START
    agents = []
    while current_row < 100:
        name = worksheet.cell(current_row, 1).value
        if name == None:
            print("Finished reading configuration !")
            break
        echo_capacity = worksheet.cell(current_row, 2).value
        irm_capacity = worksheet.cell(current_row, 3).value
        tdm_capacity = worksheet.cell(current_row, 4).value
        off = []
        for i in range(0, len(ALL_DAYSHIFTS)):
            v = worksheet.cell(current_row, EXCEL_DAYSHIFT_OFFSET + i).value
            if v == "OFF":
                off.append(ALL_DAYSHIFTS[i])
        agents.append(Agent(name, echo_capacity, irm_capacity, tdm_capacity, off))
        current_row += 1

    return Configuration(staffing_minimum, vacation_maximum, agents, demands)


if __name__ == "__main__":
    configuration = read_configuration()
    print(configuration)
